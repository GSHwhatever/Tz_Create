# -*- coding:gbk -*-
"""
��־�ҵ����̨��
"""
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime
from Reset_width import Reset
from pprint import pprint
import os, re


class JCTZ:

    def __init__(self):
        self.header = []
        self.value_lis = []
        self.sy_values = []
        self.border_style = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))  # ���õ�Ԫ��ı߿���ʽ��ȫ��Ϊϸ��
        self.alignment_style = Alignment(horizontal='center', vertical='center') # ���õ�Ԫ��Ķ��뷽ʽ�����ж���
        self.font = Font(name='����', size=10)
        self.path = os.path.join(os.path.dirname(__file__), 'template_excel')
        self.out_path = os.path.join(os.path.dirname(__file__), 'template_excel_bak')
        # self.file_tag = ''
        self.jntc = ''
        self.run_smz_status = False
        self.Rest = Reset()

    def insert_row(self, ws, data, row_index):
        ws.insert_rows(row_index)
        for col_index, value in enumerate(data, start=1):
            cell = ws.cell(row=row_index, column=col_index)
            if isinstance(value, datetime):
                cell.number_format = 'yyyy/mm/dd'
            cell.border = self.border_style
            cell.alignment = self.alignment_style
            cell.font = self.font
            cell.value = value
            
    def get_headers(self, ws, min_num, max_num):
        # ��ȡexcel��ͷ
        if min_num != max_num:
            # ���ϱ�ͷ�����
            mer_row = [mer_cell.coord for mer_cell in ws.merged_cells.ranges if (min_num in (mer_cell.min_row, mer_cell.max_row) or max_num in (mer_cell.min_row, mer_cell.max_row))]
            # �õ����ϱ�ͷ�к��кϲ���Ԫ�������
            rows_list = []
            # �Զ�������(A2,AA2,ABC2)����
            for long in range(5, 20, 2):
                rows = [i for i in mer_row if len(i)==int(long)]
                if not rows:
                    break
                # ���򣬱�֤��ͷ˳�򣬷���ֱ��д��һ������
                rows.sort(key=lambda x: x.split(':')[0])
                rows_list.extend(rows)
            header = []
            for i in rows_list:
                numbers = re.findall('[0-9]+', i)
                if numbers[0] == numbers[1]:
                    # ��Ԫ��ͬ�кϲ�
                    start, end = re.findall('[A-Z]+', i)
                    # ��������ĸת��Ϊ����
                    col_num1 = column_index_from_string(start[0])
                    col_num2 = column_index_from_string(end[0])
                    for n in range(col_num1, col_num2 + 1):
                        coord = get_column_letter(n) + str(int(i[1]) + 1)
                        if coord in [i[:2] for i in mer_row]:
                            continue
                        header.append(ws[coord].value.replace('\n', '').replace(' ', '').replace('�����', ''))

                else:
                    value = ws[i.split(':')[0]].value
                    if value:
                        header.append(value.replace('\n', '').replace(' ', '').replace('�����', ''))
            h2 = [i.value for i in ws[max_num]]
            if len(header) != len(h2):
                l = h2[len(header) - len(h2):]
                if None not in l:
                    header.extend(l)
        else:
            # ���б�ͷֱ��ȡ
            header = [i.value.replace('\n', '').replace(' ', '').replace('�����', '') for i in ws[max_num] if i.value]
        # print(header)
        return header
    
    def syry(self, ws_sy):
        header_sy = self.get_headers(ws_sy, 2, 2)
        syry_values = []
        for row in [i for i in ws_sy.iter_rows()][2:]:
            if not row[0].value:
                break
            lis_sy = []
            for i in row:
                lis_sy.append(i.value)
            syry_values.append(dict(zip(header_sy, lis_sy)))
        keys = ['���', '����', '�Ա�', '����', '�Ļ��̶�', '���֤��', '��������', '�����س�', '��ҵ��ҵ֤��', '�Ƿ�Ǽ�ʧҵ��Ա', 'ʧҵ��Ա����', 'ʧҵʱ��', '��ȡʧҵ���ս���ֹʱ��', '��ְ����', '��ѵ����', '��ҵ��������', '��ϵ�绰', '����', '�ȼ�']
        for i in syry_values:
            values = []
            for key in keys:
                v = i.get(key)
                if key == '���':
                    v = i.get('�����')
                if key == '����':
                    v = datetime.now().year - int(i.get('���֤��')[6:10])
                if key == '�Ļ��̶�':
                    v = i.get('ѧ��')
                if key == '��������':
                    v = '����'
                if key == '�����س�':
                    v = i.get('���⼼��')
                if key == '�Ƿ�Ǽ�ʧҵ��Ա':
                    v = '��'
                if key == 'ʧҵ��Ա����':
                    v = '(9)'
                if key == '��ȡʧҵ���ս���ֹʱ��':
                    v = '��'
                if key == '��ְ����':
                    v = '����'
                if key == '��ѵ����':
                    v = '��'
                if key == '��ҵ��������':
                    v = '��1��'
                if key == '��ϵ�绰':
                    v = i.get('�绰')
                if key == '����':
                    v = '�ж���'
                if key == '�ȼ�':
                    v = 'ȱ������' if i.get('���⼼��') == '��' else '�߱�����'
                if key == 'ʧҵʱ��':
                    v_time = i.get('ʧҵʱ��')
                    if v_time:
                        month = datetime.now().month - v.month
                        v = v_time.replace(month=datetime.now().month - 2) if month > 2 else v_time
                    else:
                        v = datetime.now().replace(month=datetime.now().month - 2, day=1, hour=0, minute=0, second=0, microsecond=0)
                values.append(v)
            self.sy_values.append(values)

    def read_file(self, file, min_num, max_num):
        # ��excel�ļ���ȡ���ݣ�����ͷ����������ֵ䱣����self.value_lis�б���
        # path = os.path.join(os.path.dirname(__file__), '��־�ҵ��������̨��', file)
        path = file
        # print(f'path:{path}')
        self.value_lis = []
        try:
            wb = load_workbook(path, data_only=True)
        except InvalidFileException as E:
            print(f'���ܴ���{path.split(".")[-1]}���͵�Excel�ļ�,�Ƽ����Ϊxlsx����')
            return 'error'
        else:
            if "ʵ����" in path.split('/')[-1] and 'ʧҵ��Ա���' in wb.sheetnames:
                self.syry(wb['ʧҵ��Ա���'])
            try:
                ws = wb['�¾�ҵ��Ա']
            except KeyError:
                ws = wb.worksheets[0]
            finally:
                header = self.get_headers(ws, min_num, max_num)
                for row in [i for i in ws.iter_rows()][max_num:]:
                    if not row[0].value:
                        break
                    lis = []
                    for i in row:
                        lis.append(i.value)
                    self.value_lis.append(dict(zip(header, lis)))
                # self.file_tag = file
                # pprint(self.value_lis)

    def write_excel(self, file, min_num, max_num):
        # ��self.value_lis�б��е����ݸ�����ͬ��ͷд�뵽excel�����е���Щ�����⴦��
        path = os.path.join(self.path, file)
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        mer_lis = [mer_cell.coord for mer_cell in ws.merged_cells.ranges if max_num < mer_cell.min_row or max_num < mer_cell.max_row]
        # print(f'mer_lis:{mer_lis}')
        num = max([cell.value if isinstance(cell.value, int) else 0 for cell in ws['A'][max_num:] if cell.value is not None], default=0)
        header = self.get_headers(ws, min_num, max_num)
        value_lis = []
        values_lis = self.value_lis
        if "3��ҵ������Ա����̨��" in file:
            values_lis = [i for i in self.value_lis if i.get('��ҵ������Ա') == '��']
        for n, i in enumerate(values_lis, start=1):
            # ����key�մӹ�����������
            lis = []
            for k in header:
                # �������е��ض�����
                v = i.get(k)
                if k == '���':
                    v = n + num
                
                # ̨��12
                # �Ļ��̶�
                if '12��ְ��Ա�Ǽ�̨��' in file:
                    if k == '��ѧרҵ':
                        v = '��'
                    if k == '��ְ����':
                        v = '����'
                    if k == '�ù���ʽ':
                        v = '����'
                    if k == '����н��':
                        v = '����'
                    if k == '�Ƿ�Ӧ���У��ҵ��' or k == '�Ƿ�ũ��ת���Ͷ���':
                        v = '��'
                    if k == 'ְҵָ������' or k == 'ְҵ���ܴ���':
                        v = '1'
                    if k == '�Ǽ�ʱ��':
                        v = i.get('ʧҵʱ��')
                # ̨��15
                elif '15������Ա��������������Ϣ̨��' in file:
                    if k == '����״��':
                        v = '��'
                    if k == '����Ⱥ�����':
                        pass
                    if k == 'ԭ������λ':
                        v = i.get('���ڵ�λ')
                    if k == '����ʱ��':
                        v = i.get('����/����ʱ��')
                    if k == '��������Ա��ϵ':
                        v = '����'
                    if k == '������Ա��ϵ�绰':
                        v =  i.get('��ϵ�绰')
                    if k == '���֤����':
                        v = i.get('���֤��')
                # ̨��6
                elif '6�¾�ҵ��Ա��Ϣ̨��' in file:
                    if k == '�������ƣ��С��أ�������':
                        v = '�����еε���'
                    if k == '��ҵ��λ' and not v:
                        v = i.get('��ҵ��λ(����ҵ����幤�����ݣ�')
                    if k == '��λ��ҵ':
                        v = '��' if i.get('��ҵ��ʽ') == '��λ��ҵ' else '��'
                    if k == '���幤�̻�':
                        v = '��' if i.get('��ҵ��ʽ') == '���幤�̻�' else '��'
                    if k == '�����Ը�λ':
                        v = '��' if i.get('��ҵ��ʽ') == '�����Ը�λ����' else '��'
                    if k == '����ҵ':
                        v = '��' if i.get('��ҵ��ʽ') == '����ҵ' else '��'
                    if k == 'ʧҵ��Ա�پ�ҵ':
                        # v = '��' if i.get('��ҵ����') == 'ʧҵ�پ�ҵ' else '��'
                        v = '��'
                    if k == '��ҵ�����پ�ҵ':
                        v = '��' if i.get('��ҵ������Ա') == '��' else '��'
                # ̨��5
                elif '5ʧҵ��Ա�پ�ҵ��Ϣ��ϸ̨��' in file:
                    if k == '�Ļ��̶�' and not v:
                        v = i.get('ѧ��')
                    if k == '��������':
                        v = '����'
                    if k == '�����س�' and not v:
                        v = i.get('���ܵȼ�֤��', '��')
                    if k == 'ʧҵ��Ա����':
                        v = '(9)'
                    if k == 'ʧҵʱ��':
                        d = i.get('�ǼǾ�ҵʱ�䣨��/��/�գ�')
                        if not d:
                            d = i.get('��ҵʱ��')
                        if d:
                            v = d.replace(month=d.month-1)
                    if k == '�پ�ҵʱ��' and not v:
                        v = i.get('�ǼǾ�ҵʱ�䣨��/��/�գ�')
                    if k == '��ҵ����':
                        jyqd = {"����ҵ": "(5)", "��λ��ҵ": "(3)", "���幤�̻�": "(4)", "�����Ը�λ����": "(6)"}
                        v = jyqd.get(i.get('��ҵ��ʽ'))
                    if k == '�־�ҵ��λ':
                        v = str(i.get("��ҵ��λ(����ҵ����幤�����ݣ�", "")).split("/")[0]
                    if k == '��ҵ��λ':
                        if i.get('��ҵ��ʽ') == '����ҵ':
                            v = str(i.get('��ҵ��λ(����ҵ����幤�����ݣ�', '')).split('/')[-1]
                        else:
                            v = 'Ա��'
                    if k == '������ҵ':
                        v = i.get('���²�ҵ����')
                # ̨��4
                elif '4ʧҵ��Ա����̨��' in file:
                    if k == '���':
                        v += len(self.sy_values)
                    if k == '����':
                        v = str(i.get('����', '')).split('-')[0]
                    if k == '�Ƿ�Ǽ�ʧҵ��Ա':
                        v = '��'
                    if k == '�����س�':
                        self.jnct = v = i.get('���ܵȼ�֤��')
                    if k == 'ʧҵʱ��':
                        d = i.get('�ǼǾ�ҵʱ�䣨��/��/�գ�')
                        if d:
                            v = d.replace(month=d.month-1)
                    if k == '�Ļ��̶�' and not v:
                        v = i.get('ѧ��')
                    if k == 'ʧҵ��Ա����':
                        v = '(9)'
                    if k == '��������':
                        v = '����'
                    if k == '��ȡʧҵ���ս���ֹʱ��':
                        v = '��'
                    if k == '��ְ����':
                        if i.get('��ҵ��ʽ') == '����ҵ':
                            v = str(i.get('��ҵ��λ(����ҵ����幤�����ݣ�', '')).split('/')[-1]
                        else:
                            v = 'Ա��'
                    if k == '��ѵ����':
                        v = '��'
                    if k == '��ҵ��������':
                        v = '��1��'
                    if k == '����':
                        v = '�ж���'
                    if k == '�ȼ�':
                        if '��' in self.jnct:
                            v = '�߱�����'
                        else:
                            v = 'ȱ������'
                elif "3��ҵ������Ա����̨��" in file:
                    if k == "�Ļ��̶�":
                        v = i.get('ѧ��')
                    if k == "�����س�":
                        v = i.get('���ܵȼ�֤��')
                    if k == "��ҵ������Ա����":
                        v = '��'
                    if k == "��ͥסַ":
                        v = '�����еε���'
                    if k == "�پ�ҵʱ��":
                        v = i.get('�ǼǾ�ҵʱ�䣨��/��/�գ�')
                    if k == '�־�ҵ��λ':
                        v = str(i.get("��ҵ��λ(����ҵ����幤�����ݣ�", "")).split("/")[0]
                    if k == '��ҵ��λ':
                        if i.get('��ҵ��ʽ') == '����ҵ':
                            v = str(i.get('��ҵ��λ(����ҵ����幤�����ݣ�', '')).split('/')[-1]
                        else:
                            v = 'Ա��'
                    if k == '��ҵȥ��':
                        jyqx = {"����ҵ": "��", "��λ��ҵ": "��", "���幤�̻�": "��", "�����Ը�λ����": "��"}
                        v = jyqx.get(i.get('��ҵ��ʽ'))
                    if k == '�Ƿ�ǩ���Ͷ���ͬ':
                        v = '��'
                    if k == '��ͬ����':
                        v = '��'
                    if k == '�Ƿ��ҵԮ������':
                        v = ' ��'
                    if k == '��ҵԮ����ʽ':
                        v = '��'
                lis.append(v)
            value_lis.append(lis)
        if "4ʧҵ��Ա����̨��" in file:
            value_lis = self.sy_values + value_lis
        # pprint(value_lis)
        # ��������к�
        insert_num = num + max_num
        if mer_lis:
            self.write_before(ws, mer_lis)
        for i, value in enumerate(value_lis, 1):
            self.insert_row(ws, value, insert_num + i)
        if mer_lis:
            self.write_tail(ws, mer_lis, len(value_lis))
        out_path = os.path.join(self.out_path, file)
        reset_rows = [1, 4] if "12��ְ��Ա�Ǽ�̨��" in file else [1, 4, 5]
        self.Rest.reset(ws, rows=reset_rows, value=True)
        wb.save(out_path)
        print(f'�ļ�{out_path}��д�����')

    def write_before(self, ws, mer_lis):
        for i in mer_lis:
            # num1, num2 = re.findall('[0-9]+', i)
            # ȡ����Ԫ��ϲ�
            ws.unmerge_cells(i)
        
    def write_tail(self, ws, mer_lis, rows):
        for i in mer_lis:
            num1, num2 = re.findall('[0-9]+', i)
            # ��Ԫ��ϲ�
            start, end = re.findall('[A-Z]+', i)
            ws.merge_cells(f'{start}{int(num1) + rows}:{end}{int(num2) + rows}')

    def run_smz(self, smz_path, out_path):
        # ����ʵ������Ϣ������̨��5��6
        res = self.read_file(smz_path, 2, 3)
        if res and res == 'error':
            return
        if out_path:
            self.out_path = out_path
        self.write_excel('3��ҵ������Ա����̨��.xlsx', 4, 5)
        self.write_excel('4ʧҵ��Ա����̨��.xlsx', 4, 5)
        self.write_excel('5ʧҵ��Ա�پ�ҵ��Ϣ��ϸ̨��.xlsx', 4, 5)
        self.write_excel('6�¾�ҵ��Ա��Ϣ̨��.xlsx', 4, 5)
        self.run_smz_status = True

    def run_4to12(self, tz4_path, out_path):
        # ����̨��4������̨��12
        if os.path.isfile(tz4_path):
            res = self.read_file(tz4_path, 4, 5)
        else:
            res = self.read_file(os.path.join(tz4_path, '4ʧҵ��Ա����̨��.xlsx'), 4, 5)
        if res and res == 'error':
            return
        if out_path:
            self.out_path = out_path
        self.write_excel('12��ְ��Ա�Ǽ�̨��.xlsx', 4, 4)
    
    def run_7to15(self, tz7_path, out_path):
        # ��̨��7����̨��15
        res = self.read_file(tz7_path, 4, 4)
        if res and res == 'error':
            return
        if out_path:
            self.out_path = out_path
        self.write_excel('15������Ա��������������Ϣ̨��.xlsx', 4, 5)

    def main(self):
        # ����ʵ������Ϣ������̨��3��4��5��6
        self.run_smz()

        if self.run_smz_status:
            # ����̨��4������̨��12
            self.run_4to12()

        # ��̨��7����̨��15
        # self.run_7to15()


if __name__ == '__main__':
    jctz = JCTZ()
    # jctz.run_smz(smz_path='C:/Users/Administrator/Desktop/����2024��4��ʵ����̨��20240426��.xlsx', out_path='')
    # jctz.run_smz(smz_path='C:/Users/XBD/Desktop/ʵ����.xlsx', out_path='C:/Users/XBD/Desktop/̨�˽��')
    # jctz.run_4to12(tz4_path='C:/Users/XBD/Desktop/̨�˽��/4ʧҵ��Ա����̨��.xlsx', out_path='C:/Users/XBD/Desktop/̨�˽��')
