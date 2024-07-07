# -*- coding:gbk -*-
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell import MergedCell
from icecream import ic
import datetime


class Reset:

    def __init__(self):
        self.border_style = Border(left=Side(style='thin'), 
                                    right=Side(style='thin'), 
                                    top=Side(style='thin'), 
                                    bottom=Side(style='thin'))  # ���õ�Ԫ��ı߿���ʽ��ȫ��Ϊϸ��
        self.alignment_style = Alignment(horizontal='center', vertical='center') # ���õ�Ԫ��Ķ��뷽ʽ�����ж���
        self.font = Font(name='����', size=10)
        # excel��ʽ����

    def reset(self, ws, rows=None, value=False):
        mer_lis = [mer_cell.coord.split(':')[0] for mer_cell in ws.merged_cells.ranges]
        # �޸ĵ�Ԫ����ʽ
        for row in ws:
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                coordinate = f'{cell.column_letter}{cell.row}'
                if coordinate in mer_lis:
                    continue
                if rows and cell.row in rows:
                    continue
                if value:
                    if cell.value:
                        cell.border = self.border_style
                        cell.alignment = self.alignment_style
                        cell.font = self.font
                else:
                    cell.border = self.border_style
                    cell.alignment = self.alignment_style
                    cell.font = self.font

        lks = [] # ��һ��������ÿ������ȣ����洢���б�lks�С�
        for i in range(1,ws.max_column+1): #ÿ��ѭ��
            lk = 1 #�����ʼ�п�����ÿ����ѭ����ɺ�����
            for j in range(1, ws.max_row + 1): #ÿ��ѭ��
                cell = ws.cell(row=j,column=i)
                if isinstance(cell, MergedCell):
                    continue
                coordinate = f'{cell.column_letter}{cell.row}'
                if coordinate in mer_lis:
                    continue
                if rows and cell.row in rows:
                    continue
                sz = cell.value #ÿ����Ԫ������
                if isinstance(sz,str): #
                    try:
                        lk1 = len(sz.encode('gbk')) #gbk����һ���������ֽڣ�utf-8һ���������ֽڣ�gbk����
                    except UnicodeEncodeError:
                        lk1 = len(sz.encode('utf-8')) - 1
                elif isinstance(sz, datetime.datetime):
                    lk1 = len(sz.strftime("%Y/%m/%d"))
                else:
                    lk1 = len(str(sz))
                if lk < lk1:
                    lk = lk1 #����ÿ��ѭ�������ֵ����lk��
                # ic(lk)
            lks.append(lk) # ��ÿ������ȼ����б�������һ������lks = lks.append(lk)����append���޸��б����������ֵnone����none���ܼ�����append������
        
        # �ڶ����������п�
        for i in range(1, ws.max_column +1):
            k = get_column_letter(i) #������ת��Ϊ����,26����ĸ����Ҳ������[chr(i).upper() for i in range(97, 123)]�����õ���ģ��\
            ws.column_dimensions[k].width = lks[i-1] + 2 #�����п�һ��������ֽڿ�ȣ����Ը���ʵ�����������


if __name__ == "__main__":
    R = Reset()
    path = r'C:\Users\GSH\Desktop\4ʧҵ��Ա����̨�� - ����.xlsx'
    wb = load_workbook(path)
    ws = wb.active
    R.reset(ws)
    wb.save(path)
