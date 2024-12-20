# -*- coding:gbk -*-
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from collections import Counter
from datetime import datetime
import os


class Write:

    def __init__(self) -> None:
        self.smz_path = r'C:/Users/GSH\Desktop/小半道社区2024实名制.xlsx'
        out_path = os.path.join(os.environ['USERPROFILE'], 'Desktop', '台账结果', '就业统计报表.xlsx')
        self.bb_path = out_path if os.path.exists(out_path) else os.path.join(os.path.dirname(__file__), 'template_excel', '就业统计报表.xlsx')
        self.out_path = os.path.join(os.environ['USERPROFILE'], 'Desktop', '就业统计报表.xlsx')
        self.industries = [
            "农、林、牧、渔业",
            "采矿业",
            "制造业",
            "电力、燃气及水的生产和供应业",
            "建筑业",
            "交通运输、仓储和邮政业",
            "信息传输、计算机服务和软件业",
            "批发和零售业",
            "住宿和餐饮业",
            "金融业",
            "房地产业",
            "租赁和商务服务业",
            "科学研究、技术服务和地质勘查业",
            "水利、环境和公共设施管理业",
            "居民服务和其他服务业",
            "教育",
            "卫生、社会保障和社会福利业",
            "文化、体育和娱乐业",
            "冰雪旅游业",
            "公共管理和社会组织"
        ]
        self.EP1 = []
        self.sjy01= []
        self.sy02 = []
        self.hyhf = []
    
    def re_init(self):
        self.EP1 = []
        self.sjy01 = []
        self.sy02 = []
        self.hyhf = []
    
    def read(self, month : int):
        smz_wb = load_workbook(self.smz_path)
        all_ws = smz_wb['新就业人员']
        xzjy_ws = smz_wb['新增就业人员']
        zrjy_ws = smz_wb['自然减员（就业）']
        syry_ws = smz_wb['失业人员情况']

        # print([v.value for i, v in zip(zrjy_ws['A'], zrjy_ws['I']) if (i.value and isinstance(v.value, datetime)) and v.value.month <= datetime.now().month])
        xzjy_num = len([i for i, v in zip(xzjy_ws['A'], xzjy_ws['J'])  if (i.value and isinstance(v.value, datetime)) and v.value.month <= month])        # 新增就业人数
        zrjy_num = len([i for i, v in zip(zrjy_ws['A'], zrjy_ws['I']) if (i.value and isinstance(v.value, datetime)) and v.value.month <= month])        # 自然减员人数
        
        syzjy_num = len([i for i, v in zip(all_ws['K'], all_ws['H']) if i.value and i.value == '失业再就业' and isinstance(v.value, datetime) and v.value.month <= month])        # 失业再就业人数
        jykn_num = len([i for i, v in zip(all_ws['R'], all_ws['H']) if i.value and i.value == '是' and isinstance(v.value, datetime) and v.value.month <= month])     # 就业困难人数

        xl_num = len([i for i, v in zip(all_ws['G'], all_ws['H']) if i.value and i.value in ('大学专科', '大学本科', '硕士研究生') and isinstance(v.value, datetime) and v.value.month <= month])     # 大专以上学历人数
        nv_num = len([i for i, v in zip(all_ws['X'], all_ws['H']) if i.value and i.value == '女' and isinstance(v.value, datetime) and v.value.month <= month])       # 女性人数
        cylb_dic = Counter([i.value for i, v in zip(all_ws['P'], all_ws['H']) if i.value and i.value in ('第一产业', '第二产业', '第三产业') and isinstance(v.value, datetime) and v.value.month <= month])     # 产业类别计数
        jyqd_dic = Counter([i.value for i, v in zip(all_ws['J'], all_ws['H']) if i.value and i.value in ('单位就业', '灵活就业', '个体工商户', '公益性岗位安置') and isinstance(v.value, datetime) and v.value.month <= month])     # 就业渠道计数
        xydm_lis = [i.value for i, j, v in zip(all_ws['L'], all_ws['J'], all_ws['H']) if j == '单位就业' and i.value and i.value.isalnum() and isinstance(v.value, datetime) and v.value.month <= month]      # 单位就业人员单位统一就业代码
        sydw_num = len([i for i in xydm_lis if i[0] == '1'])      # 根据统一信用代码筛选第一位为'1'的为机关事业单位
        nlhf_dic = Counter([i.value for i, v in zip(all_ws['AF'], all_ws['H']) if i.value and i.value in ('16-24', '25-45', '46-60') and isinstance(v.value, datetime) and v.value.month <= month])     # 年龄划分计数

        syry_num = len([i for i in syry_ws['A'] if i.value]) - 2        # 失业人员人数
        synx_num = len([i for i in syry_ws['C'] if i.value and i.value == '女'])        # 失业人员其中女性人数
        sykn_num = len([i for i in syry_ws['K'] if i.value and i.value == '就业困难'])        # 失业人员其中就业困难人数

        hyhf_dic = Counter([i.value for i, v in zip(all_ws['N'], all_ws['H']) if i.value and i.value in self.industries and isinstance(v.value, datetime) and v.value.month <= month])
        
        self.EP1.extend([xzjy_num, xzjy_num + zrjy_num, zrjy_num, syzjy_num, jykn_num])
        self.sjy01.extend([xzjy_num + zrjy_num, syzjy_num, jykn_num, xl_num, nv_num, cylb_dic.get('第一产业', 0), cylb_dic.get('第二产业', 0), cylb_dic.get('第三产业', 0),
                           0, 0, jyqd_dic.get('单位就业', 0) - sydw_num, sydw_num, jyqd_dic.get('个体工商户', 0), jyqd_dic.get('灵活就业', 0), jyqd_dic.get('公益性岗位安置', 0), nlhf_dic.get('16-24', 0),
                           nlhf_dic.get('25-45', 0), nlhf_dic.get('46-60', 0)])
        self.sy02.extend([syry_num, 0, 0, 0, 0, 0, syry_num, 0, 0, 0, 0, 0, synx_num, sykn_num])
        self.hyhf.extend([xzjy_num + zrjy_num])
        self.hyhf.extend([hyhf_dic.get(i, 0) for i in self.industries])
        # print(self.EP1)
        # print(self.sjy01)
        # print(self.sy02)
        # print(self.hyhf)
    
    def write(self, month : int, bb_wb):
        row = month + 5
        for sheet, datas in [('人社统EP1', self.EP1), ('省就业01', self.sjy01), ('02城镇表登记失业人员情况', self.sy02), ('行业划分', self.hyhf)]:
            ws = bb_wb[sheet]
            for i, v in enumerate(datas, start=3):
                ws.cell(row=row, column=i, value=v)

    def run(self, path, out_path):
        self.smz_path = path
        self.out_path = os.path.join(out_path, '就业统计报表.xlsx')
        smz_wb = load_workbook(self.smz_path)
        all_ws = {i.value.month: '' for i in smz_wb['新就业人员']['H'] if i.value and isinstance(i.value, datetime)}
        bb_wb = load_workbook(self.bb_path)
        for month in list(all_ws.keys()):
            self.re_init()
            self.read(month)
            self.write(month, bb_wb)
        bb_wb.save(self.out_path)
        print(f'文件{self.out_path}，写入完成')
    
    def main(self):
        smz_wb = load_workbook(self.smz_path)
        all_ws = {i.value.month: '' for i in smz_wb['新就业人员']['H'] if i.value and isinstance(i.value, datetime)}
        bb_wb = load_workbook(self.bb_path)
        for month in list(all_ws.keys()):
            self.re_init()
            self.read(month)
            self.write(month, bb_wb)
        bb_wb.save(self.out_path)
        print(f'文件{self.out_path}，写入完成')


if __name__ == '__main__':
    W = Write()
    W.main()
    


        

